#!/usr/bin/env python3
"""Schreibt Embeddings von Abstracts (Spalte D) in Spalte G einer Excel-Datei.

Beispiel:
    python excel_ollama_embeddings.py \
        --input papers.xlsx \
        --model nomic-embed-text \
        --sheet "Sheet1"
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl
import requests
from openpyxl.utils.cell import column_index_from_string, get_column_letter


class OllamaClient:
    def __init__(self, base_url: str, model: str, timeout: int = 120) -> None:
        self.base_url = base_url.rstrip("/")
        self.model = model
        self.timeout = timeout

    def embed(self, text: str) -> list[float]:
        """Erzeuge Embedding für einen Text. Unterstützt /api/embed und /api/embeddings."""
        payload_embed = {"model": self.model, "input": text}
        payload_embeddings = {"model": self.model, "prompt": text}

        # Neue API
        url_embed = f"{self.base_url}/api/embed"
        resp = requests.post(url_embed, json=payload_embed, timeout=self.timeout)
        if resp.ok:
            data = resp.json()
            emb = data.get("embeddings")
            if isinstance(emb, list) and emb:
                if isinstance(emb[0], list):
                    return emb[0]
                if isinstance(emb[0], (int, float)):
                    return emb
            raise ValueError(f"Unerwartete Antwort von {url_embed}: {data}")

        # Fallback: alte API
        url_embeddings = f"{self.base_url}/api/embeddings"
        resp2 = requests.post(url_embeddings, json=payload_embeddings, timeout=self.timeout)
        if resp2.ok:
            data2 = resp2.json()
            emb2 = data2.get("embedding")
            if isinstance(emb2, list):
                return emb2
            raise ValueError(f"Unerwartete Antwort von {url_embeddings}: {data2}")

        raise RuntimeError(
            "Ollama-Embedding fehlgeschlagen. "
            f"/api/embed: HTTP {resp.status_code}, /api/embeddings: HTTP {resp2.status_code}."
        )


def process_workbook(
    input_path: Path,
    output_path: Path,
    client: OllamaClient,
    sheet_name: str | None,
    start_row: int,
    abstract_col: str,
    output_col: str,
    output_mode: str,
    output_prefix: str,
    write_header: bool,
) -> tuple[int, int]:
    wb = openpyxl.load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    if write_header and start_row <= 1 and output_mode == "json":
        ws[f"{output_col}1"] = "Embedding"

    processed = 0
    skipped = 0

    max_row = ws.max_row
    for row in range(start_row, max_row + 1):
        abstract_value = ws[f"{abstract_col}{row}"].value
        if abstract_value is None:
            skipped += 1
            continue

        text = str(abstract_value).strip()
        if not text:
            skipped += 1
            continue

        embedding = client.embed(text)
        if output_mode == "json":
            ws[f"{output_col}{row}"] = json.dumps(embedding, ensure_ascii=False)
        else:
            start_idx = column_index_from_string(output_col)
            for i, value in enumerate(embedding):
                col_idx = start_idx + i
                col_letter = get_column_letter(col_idx)
                ws[f"{col_letter}{row}"] = float(value)
                if write_header and start_row <= 1:
                    ws[f"{col_letter}1"] = f"{output_prefix}{i + 1}"
        processed += 1
        print(f"Zeile {row}: Embedding geschrieben (Länge={len(embedding)})")

    wb.save(output_path)
    return processed, skipped


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Erzeuge Ollama-Embeddings für Abstracts aus Excel (D -> G)."
    )
    parser.add_argument("--input", required=True, type=Path, help="Eingabe-xlsx")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Ausgabe-xlsx (Standard: <input>_with_embeddings.xlsx)",
    )
    parser.add_argument("--model", required=True, help="Ollama Embedding-Modell")
    parser.add_argument(
        "--ollama-url",
        default="http://localhost:11434",
        help="Basis-URL des lokalen Ollama-Servers",
    )
    parser.add_argument("--sheet", default=None, help="Worksheet-Name (optional)")
    parser.add_argument("--start-row", type=int, default=2, help="Startzeile für Daten")
    parser.add_argument("--abstract-col", default="D", help="Spalte mit Abstract")
    parser.add_argument("--output-col", default="G", help="Spalte für Embedding")
    parser.add_argument(
        "--output-mode",
        choices=["json", "columns"],
        default="json",
        help="json: eine Zelle pro Embedding; columns: eine Dimension pro Spalte (MATLAB-freundlich)",
    )
    parser.add_argument(
        "--output-prefix",
        default="emb_",
        help="Spalten-Präfix für --output-mode columns (z. B. emb_ -> emb_1, emb_2, ...)",
    )
    parser.add_argument(
        "--no-header",
        action="store_true",
        help="Schreibt keinen Header in der Ausgabespalte",
    )
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)

    if not args.input.exists():
        print(f"Fehler: Datei nicht gefunden: {args.input}", file=sys.stderr)
        return 1

    output = args.output
    if output is None:
        output = args.input.with_name(f"{args.input.stem}_with_embeddings{args.input.suffix}")

    client = OllamaClient(base_url=args.ollama_url, model=args.model)

    try:
        processed, skipped = process_workbook(
            input_path=args.input,
            output_path=output,
            client=client,
            sheet_name=args.sheet,
            start_row=args.start_row,
            abstract_col=args.abstract_col,
            output_col=args.output_col,
            output_mode=args.output_mode,
            output_prefix=args.output_prefix,
            write_header=not args.no_header,
        )
    except Exception as exc:  # noqa: BLE001
        print(f"Fehler bei Verarbeitung: {exc}", file=sys.stderr)
        return 1

    print(
        f"Fertig. Verarbeitet: {processed}, Übersprungen: {skipped}, Ausgabe: {output}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
